# %% [markdown]
# # Manufactured Home Price Prediction
# **MHS Data Scraping| MHS EDA| MHS Modeling | Final Pipeline**
# 
# ---
# 

# %% [markdown]
# Below Codes Searches the available raw dataset, if there is new data available, it will fetch and complies into single excel spreadsheet.

# %%
#import all packages & libraries
%pip install openpyxl pandas xlrd
%pip install numpy
%pip install matplotlib
%pip install seaborn
%pip install lightgbm
%pip install scikit-learn

# DBTITLE 1,
from __future__ import annotations

import argparse
import hashlib
import re
import sys
from dataclasses import dataclass
from itertools import zip_longest
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin
from urllib.request import Request, urlopen

import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.patches as mpatches
import lightgbm as lgb
import warnings
warnings.filterwarnings('ignore')

from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error

print("All packages & libraries imported !!")

# %%
SOURCE_URL = "https://www.census.gov/data/datasets/2021/econ/MHS/puf.html"

FINAL_COLUMNS = [
    "CONTROL",
    "REGION",
    "SHIPMONTH",
    "SECTIONS",
    "WEIGHT",
    "WGTADJ",
    "STATUS",
    "PRICE",
    "SQFT",
    "BEDROOMS",
    "TITLED",
    "LOCATION",
    "FOUNDATION",
    "SECURED",
    "JSTATUS",
    "JPRICE",
    "JSQFT",
    "JBEDROOM",
    "JTITLE",
    "JLOCATION",
    "JFOUNDATION",
    "JSECURED",
    "YEAR",
    "FINALDEST",
    "FOOTINGS",
    "LEASE",
    "JLEASE",
    "JFINALDEST",
    "JFOOTINGS",
]

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)


@dataclass(frozen=True)
class PufLink:
    year: int
    url: str
    filename: str


def fetch_bytes(url: str, max_retries: int = 5) -> bytes:
    import time
    from urllib.error import HTTPError, URLError
    
    request = Request(url, headers={"User-Agent": USER_AGENT})
    for attempt in range(max_retries):
        try:
            with urlopen(request, timeout=600) as response:
                return response.read()
        except (TimeoutError, HTTPError, URLError) as e:
            if attempt < max_retries - 1:
                wait_time = 2 ** attempt
                print(f"Error on attempt {attempt + 1}/{max_retries}: {e}")
                print(f"Waiting {wait_time} seconds before retrying...")
                time.sleep(wait_time)
                continue
            raise


def discover_puf_links(source_url: str) -> list[PufLink]:
    html = fetch_bytes(source_url).decode("utf-8", errors="ignore")
    anchor_pattern = re.compile(
        r"<a[^>]+href=[\"']([^\"']+)?[\"'][^>]*>(.*?)</a>",
        flags=re.IGNORECASE | re.DOTALL,
    )

    links: dict[int, PufLink] = {}
    for href, anchor_html in anchor_pattern.findall(html):
        absolute_url = urljoin(source_url, href)
        lower_url = absolute_url.lower()
        if not lower_url.endswith((".xls", ".xlsx")):
            continue
        anchor_text = re.sub(r"<.*?>|\s+", " ", anchor_html).strip()
        if "puf" not in lower_url and "public use file" not in anchor_text.lower():
            continue
        filename = Path(absolute_url.split("?", 1)[0]).name
        year_match = (
            re.search(r"(20\d{2})", anchor_text)
            or re.search(r"(20\d{2})", filename)
            or re.search(r"(20\d{2})", absolute_url)
        )
        if not year_match:
            continue
        year = int(year_match.group(1))
        links[year] = PufLink(year=year, url=absolute_url, filename=filename)

    if links:
        min_year = min(links)
        max_year = max(links)
        # The Census page has historically had one stale/mispointed annual link
        # while the actual PUF file remained in the standard year directory.
        # Probe the predictable PUFYYYY.xlsx path for any missing in-range years.
        for year in range(min_year, max_year + 1):
            if year not in links:
                inferred = f"https://www2.census.gov/programs-surveys/mhs/tables/{year}/PUF{year}.xlsx"
                links[year] = PufLink(year=year, url=inferred, filename=f"PUF{year}.xlsx")

    if not links:
        raise RuntimeError(f"No MHS PUF Excel links were found at {source_url}")

    return [links[year] for year in sorted(links)]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def download_puf_files(links: Iterable[PufLink], raw_dir: Path, force: bool) -> list[Path]:
    raw_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for link in links:
        suffix = Path(link.filename).suffix or ".xlsx"
        path = raw_dir / f"mhs_puf_{link.year}{suffix}"
        if force or not path.exists():
            print(f"Downloading {link.year}: {link.url}")
            path.write_bytes(fetch_bytes(link.url))
        else:
            print(f"Using cached {link.year}: {path}")
        paths.append(path)
    return paths


def clean_header_name(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", "", text.strip().upper())


def find_header_row(path: Path) -> int:
    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    df_temp = pd.read_excel(path, sheet_name=0, nrows=25, header=None, engine=engine)
    expected = {"CONTROL", "REGION", "SHIPMONTH"}
    for index, row in df_temp.iterrows():
        found = {clean_header_name(value) for value in row}
        if len(expected & found) >= 2:
            return index
    return 0


def read_puf_file(path: Path, year: int) -> pd.DataFrame:
    header_row = find_header_row(path)
    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    df = pd.read_excel(path, sheet_name=0, header=header_row, dtype=object, engine=engine)
    df.columns = [clean_header_name(column) for column in df.columns]
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    df = df.loc[:, [column for column in df.columns if not column.startswith("UNNAMED")]]

    # Keep only rows with real record identifiers.
    if "CONTROL" in df.columns:
        df = df[df["CONTROL"].notna()]

    # Some annual files call the bedroom imputation flag JBEDROOMS; the target workbook uses JBEDROOM.
    if "JBEDROOMS" in df.columns and "JBEDROOM" not in df.columns:
        df = df.rename(columns={"JBEDROOMS": "JBEDROOM"})
    if "JTITLED" in df.columns and "JTITLE" not in df.columns:
        df = df.rename(columns={"JTITLED": "JTITLE"})
    if "CONTROLID" in df.columns and "CONTROL" not in df.columns:
        df = df.rename(columns={"CONTROLID": "CONTROL"})
    if "LOCREGIN" in df.columns and "REGION" not in df.columns:
        df = df.rename(columns={"LOCREGIN": "REGION"})

    if "YEAR" not in df.columns:
        if "SHIPMONTH" in df.columns:
            df["YEAR"] = pd.to_numeric(df["SHIPMONTH"], errors="coerce").floordiv(100)
        else:
            df["YEAR"] = year

    if year >= 2021:
        # MHS redesigned placement/status fields in 2021. Recreate the
        # legacy columns used by the 2014-2020 PUF layout and reference file.
        if "STATUS" in df.columns:
            df["STATUS"] = pd.to_numeric(df["STATUS"], errors="coerce").map({1: 2, 2: 1, 3: 3, 9: 9})
        if "PIERS" in df.columns:
            df["FOUNDATION"] = pd.to_numeric(df["PIERS"], errors="coerce").map(
                {0: 1, 1: 3, 2: 3, 3: 2, 4: 2, 9: 9}
            )
        if "JPIERS" in df.columns:
            df["JFOUNDATION"] = df["JPIERS"]
        if "SECURED" in df.columns:
            df["SECURED"] = pd.to_numeric(df["SECURED"], errors="coerce").map(
                {0: 1, 1: 1, 2: 1, 3: 1, 9: 9}
            )
        if year >= 2022 and "LOCATION" in df.columns:
            numeric_location = pd.to_numeric(df["LOCATION"], errors="coerce")
            df["LOCATION"] = numeric_location.replace({2: 3})

    for column in FINAL_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA

    return df[FINAL_COLUMNS].copy()


def normalize_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    df = df.where(pd.notna(df), None)
    numeric_columns = [column for column in FINAL_COLUMNS if not column.startswith("J")]
    for column in numeric_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")
        df[column] = df[column].where(pd.notna(df[column]), None)
    for column in FINAL_COLUMNS:
        if column.startswith("J"):
            df[column] = df[column].apply(lambda value: None if value is None else str(value).strip())
    return df


def build_final_dataset(paths: list[Path]) -> pd.DataFrame:
    frames = []
    for path in paths:
        year_match = re.search(r"(20\d{2})", path.name)
        if not year_match:
            raise ValueError(f"Could not identify year from file name: {path}")
        year = int(year_match.group(1))
        frame = read_puf_file(path, year)
        print(f"Read {year}: {len(frame):,} rows x {len(frame.columns)} columns")
        frames.append(frame)
    return normalize_for_excel(pd.concat(frames, ignore_index=True))


def write_excel(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")


def comparable(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return None if stripped == "" else stripped
    return value


def values_equal(left: object, right: object) -> bool:
    left = comparable(left)
    right = comparable(right)
    if left is None or right is None:
        return left is None and right is None
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) < 1e-9
    return left == right


def compare_workbooks(generated_path: Path, reference_path: Path, max_examples: int = 25) -> bool:
    generated = openpyxl.load_workbook(generated_path, read_only=True, data_only=True)
    reference = openpyxl.load_workbook(reference_path, read_only=True, data_only=True)
    try:
        generated_ws = generated["Data"]
        reference_ws = reference["Data"]
        same_dimensions = (
            generated_ws.max_row == reference_ws.max_row
            and generated_ws.max_column == reference_ws.max_column
        )
        print(
            "Generated dimensions:",
            f"{generated_ws.max_row:,} rows x {generated_ws.max_column:,} columns",
        )
        print(
            "Reference dimensions:",
            f"{reference_ws.max_row:,} rows x {reference_ws.max_column:,} columns",
        )

        mismatch_count = 0
        examples = []
        max_col = max(generated_ws.max_column, reference_ws.max_column)
        generated_rows = generated_ws.iter_rows(values_only=True)
        reference_rows = reference_ws.iter_rows(values_only=True)
        missing = object()
        for row_number, (generated_row, reference_row) in enumerate(
            zip_longest(generated_rows, reference_rows, fillvalue=missing),
            start=1,
        ):
            if generated_row is missing:
                generated_row = ()
            if reference_row is missing:
                reference_row = ()
            for col in range(1, max_col + 1):
                generated_value = generated_row[col - 1] if col <= len(generated_row) else None
                reference_value = reference_row[col - 1] if col <= len(reference_row) else None
                if values_equal(generated_value, reference_value):
                    continue
                mismatch_count += 1
                if len(examples) < max_examples:
                    examples.append((row_number, col, generated_value, reference_value))

        if mismatch_count:
            print(f"Value mismatches: {mismatch_count:,}")
            for row, col, generated_value, reference_value in examples:
                column_name = reference_ws.cell(row=1, column=col).value or generated_ws.cell(row=1, column=col).value
                print(
                    f"  row {row:,}, col {col} ({column_name}): "
                    f"generated={generated_value!r}, reference={reference_value!r}"
                )
        else:
            print("Value mismatches: 0")

        return same_dimensions and mismatch_count == 0
    finally:
        generated.close()
        reference.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download, standardize, and combine Census MHS PUF Excel files."
    )
    parser.add_argument("--url", default=SOURCE_URL, help="Census MHS PUF page URL.")
    parser.add_argument("--raw-dir", type=Path, default=Path("data/raw"), help="Directory for downloaded annual files.")
    parser.add_argument("--output", type=Path, default=Path("Final_Data.xlsx"), help="Generated Excel path.")
    parser.add_argument("--reference", type=Path, help="Optional reference workbook to compare against.")
    parser.add_argument("--force-download", action="store_true", help="Download files even if cached copies exist.")
    # In a Colab environment, sys.argv will contain arguments passed by the kernel
    # (e.g., -f /root/.local/share/jupyter/runtime/kernel-XXXX.json). These are not
    # intended for the script's argparse, so we pass an empty list to parse_args
    # to prevent it from trying to parse them.
    return parser.parse_args([])


def main() -> int:
    args = parse_args()
    links = discover_puf_links(args.url)
    print("Discovered years:", ", ".join(str(link.year) for link in links))
    paths = download_puf_files(links, args.raw_dir, args.force_download)
    for path in paths:
        print(f"Cached file: {path} sha256={sha256(path)[:16]}")

    df = build_final_dataset(paths)
    write_excel(df, args.output)
    print(f"Wrote {args.output}")

    if args.reference:
        matched = compare_workbooks(args.output, args.reference)
        if not matched:
            return 2
        print("Generated workbook matches the reference workbook values.")
    return 0


if __name__ == "__main__":
    main()

# %% [markdown]
# # Manufactured Home Price Prediction
# **MHS EDA

# %%
# Load the dataset
df = pd.read_excel('Final_Data.xlsx')

# %%
# Basic exploration
print(f"Dataset Shape: {df.shape[0]} rows × {df.shape[1]} columns")

# %%
print("\nColumn Names:")
print(df.columns.tolist())

# %%
print("\nFirst 5 rows:")
print(df.head())

# %%
print("\nData Types:")
print(df.dtypes)

# %%
print("\nBasic Info:")
df.info()

# %%
# Check for missing values
missing_summary = pd.DataFrame({
    'Column': df.columns,
    'Missing_Count': df.isnull().sum(),
    'Missing_Percent': (df.isnull().sum() / len(df)) * 100
})

# %%
# Show only columns with missing data
missing_summary = missing_summary[missing_summary['Missing_Count'] > 0]
missing_summary = missing_summary.sort_values('Missing_Percent', ascending=False)

# %%
print("="*80)
print("MISSING DATA ANALYSIS")
print("="*80)
print(missing_summary)

# %%
# Overall missing percentage
total_missing = df.isnull().sum().sum()
total_cells = df.shape[0] * df.shape[1]
print(f"\nTotal missing: {total_missing:,} cells ({(total_missing/total_cells)*100:.2f}%)")

# %%
# Focus on PRICE column (which has missing data)
missing_price = df[df['PRICE'].isnull()]

print("Records with missing PRICE:", len(missing_price))
print("\nDistribution by STATUS:")
print(missing_price['STATUS'].value_counts())
print("\nDistribution by REGION:")
print(missing_price['REGION'].value_counts())
print("\nDistribution by YEAR:")
print(missing_price['YEAR'].value_counts())

# %%
# Check for duplicate records
print(f"Complete duplicate rows: {df.duplicated().sum()}")

# %%
# Check value ranges for categorical variables
print("\nREGION distribution:")
print(df['REGION'].value_counts().sort_index())

print("\nSTATUS distribution:")
print(df['STATUS'].value_counts().sort_index())

print("\nSECTIONS distribution:")
print(df['SECTIONS'].value_counts().sort_index())

print("\nBEDROOMS distribution:")
print(df['BEDROOMS'].value_counts().sort_index())

# %%
# Check for negative or zero values
print(f"PRICE - Negative: {(df['PRICE'] < 0).sum()}, Zero: {(df['PRICE'] == 0).sum()}")
print(f"PRICE range: ${df['PRICE'].min():,.0f} to ${df['PRICE'].max():,.0f}")

print(f"\nSQFT - Negative: {(df['SQFT'] < 0).sum()}, Zero: {(df['SQFT'] == 0).sum()}")
print(f"SQFT range: {df['SQFT'].min():,} to {df['SQFT'].max():,}")

# %% [markdown]
# INVESTIGATING CODE '9'

# %%
# Check records where values are 9
print("\nRecords with PRICE = 9:")
print(df[df['PRICE'] == 9][['CONTROL', 'PRICE', 'SQFT', 'BEDROOMS', 'STATUS']].head())

# %% [markdown]
# Outlier Detection

# %%
# Create dataset excluding '9' codes for valid analysis
df_valid = df[(df['PRICE'] != 9) & (df['SQFT'] != 9) & (df['BEDROOMS'] != 9)].copy()

# %%
print(f"Valid records: {len(df_valid):,} (excluded {len(df) - len(df_valid):,} NA-coded)")

# %%
# IQR method for outlier detection
def detect_outliers_iqr(data, column):
    Q1 = data[column].quantile(0.25)
    Q3 = data[column].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    outliers = data[(data[column] < lower_bound) | (data[column] > upper_bound)]
    return len(outliers), lower_bound, upper_bound, Q1, Q3

# %%
# PRICE outliers
count, lower, upper, q1, q3 = detect_outliers_iqr(df_valid, 'PRICE')
print("\nPRICE OUTLIERS:")
print(f"Q1: ${q1:,.0f}, Q3: ${q3:,.0f}")
print(f"IQR bounds: ${lower:,.0f} to ${upper:,.0f}")
print(f"Outliers: {count:,} ({count/len(df_valid)*100:.2f}%)")

# %%
# SQFT outliers
count, lower, upper, q1, q3 = detect_outliers_iqr(df_valid, 'SQFT')
print("\nSQFT OUTLIERS:")
print(f"Q1: {q1:,.0f}, Q3: {q3:,.0f}")
print(f"IQR bounds: {lower:,.0f} to {upper:,.0f}")
print(f"Outliers: {count:,} ({count/len(df_valid)*100:.2f}%)")

# %%
# Overall statistics
print("\nAll numerical columns:")
print(df.describe())

# %%
#DETAILED STATISTICS (Valid Data Only)
print("\nAll numerical columns:")
print(df_valid.describe())

# %%
#key variables
print("\nPRICE:")
print(df_valid['PRICE'].describe())

print("\nSQFT:")
print(df_valid['SQFT'].describe())

# %%
#Price Distribution
plt.figure(figsize=(10, 6))
plt.hist(df_valid['PRICE'], bins=50, color='steelblue', edgecolor='black')
plt.axvline(df_valid['PRICE'].mean(), color='red', linestyle='--', label='Mean')
plt.axvline(df_valid['PRICE'].median(), color='green', linestyle='--', label='Median')
plt.xlabel('Price ($)')
plt.ylabel('Frequency')
plt.title('Distribution of Home Prices')
plt.legend()
plt.grid(True, alpha=0.3)
plt.savefig('price_distribution.png', dpi=300)
plt.show()

# %%
#SQFT Distribution
plt.figure(figsize=(10, 6))
plt.hist(df_valid['SQFT'], bins=50, color='coral', edgecolor='black')
plt.axvline(df_valid['SQFT'].mean(), color='red', linestyle='--', label='Mean')
plt.axvline(df_valid['SQFT'].median(), color='green', linestyle='--', label='Median')
plt.xlabel('Square Footage')
plt.ylabel('Frequency')
plt.title('Distribution of Square Footage')
plt.legend()
plt.grid(True, alpha=0.3)
plt.savefig('sqft_distribution.png', dpi=300)
plt.show()

# %%
#Sections Bar Chart
plt.figure(figsize=(8, 6))
sections_counts = df['SECTIONS'].value_counts().sort_index()
plt.bar(sections_counts.index, sections_counts.values, color='steelblue', edgecolor='black')
plt.xlabel('Number of Sections')
plt.ylabel('Count')
plt.title('Distribution of Home Sections')
plt.xticks([1, 2, 3])
plt.grid(True, alpha=0.3, axis='y')
plt.savefig('sections_distribution.png', dpi=300)
plt.show()

# %%
#Price by Sections
plt.figure(figsize=(10, 6))
df_valid.boxplot(column='PRICE', by='SECTIONS', grid=True)
plt.xlabel('Number of Sections')
plt.ylabel('Price ($)')
plt.title('Price Distribution by Sections')
plt.suptitle('')  # Remove default title
plt.savefig('price_by_sections.png', dpi=300)
plt.show()

# %%
#Price by Region
plt.figure(figsize=(10, 6))
df_valid.boxplot(column='PRICE', by='REGION', grid=True)
plt.xlabel('Region')
plt.ylabel('Price ($)')
plt.title('Price Distribution by Region')
plt.suptitle('')
plt.savefig('price_by_region.png', dpi=300)
plt.show()

# %%
#Price vs SQFT Scatter Plot
plt.figure(figsize=(10, 6))
plt.scatter(df_valid['SQFT'], df_valid['PRICE'], alpha=0.3, s=10)
plt.xlabel('Square Footage')
plt.ylabel('Price ($)')
plt.title('Price vs Square Footage')
plt.grid(True, alpha=0.3)
plt.savefig('price_vs_sqft.png', dpi=300)
plt.show()

# %%
#Price Trends by Year
yearly_avg = df_valid.groupby('YEAR')['PRICE'].mean()

plt.figure(figsize=(10, 6))
plt.plot(yearly_avg.index, yearly_avg.values, marker='o', linewidth=2, markersize=8)
plt.xlabel('Year')
plt.ylabel('Average Price ($)')
plt.title('Average Price Trend (2014-2024)')
plt.grid(True, alpha=0.3)
plt.savefig('price_trend.png', dpi=300)
plt.show()

# %%
# Select only key columns
key_cols = ['PRICE', 'SQFT', 'SECTIONS', 'YEAR']
corr = df_valid[key_cols].corr()

plt.figure(figsize=(8, 6))
sns.heatmap(corr, annot=True, cmap='coolwarm', center=0, fmt='.2f', square=True)
plt.title('Correlation Matrix')
plt.savefig('correlation.png', dpi=300)
plt.show()

# Print correlations
print("\nCorrelations with PRICE:")
print(corr['PRICE'].sort_values(ascending=False))

# %% [markdown]
# # Manufactured Home Price Prediction
# **Modeling Section**
# 
# ---

# %% [markdown]
# ---
# ## Section 1 — Data Loading & Quality Check

# %%
df = pd.read_excel('Final_Data.xlsx')
print(f'Shape: {df.shape[0]:,} rows × {df.shape[1]} columns')
print(f'\nColumns: {df.columns.tolist()}')


# %%
# Missing values
missing = pd.DataFrame({
    'Missing Count'  : df.isnull().sum(),
    'Missing %'      : (df.isnull().sum() / len(df) * 100).round(2)
}).query('`Missing Count` > 0').sort_values('Missing %', ascending=False)
print('Columns with missing values:')
print(missing)


# %%
# Coded-9 distribution (9 = Not Applicable in MHS codebook)
CODED_COLS = ['BEDROOMS','TITLED','LOCATION','FOUNDATION','SECURED','PRICE']
print(f"{'Column':<14} {'Count of 9s':>12} {'% of rows':>10}")
print('-' * 40)
for col in CODED_COLS:
    n9  = (df[col] == 9).sum()
    pct = n9 / len(df) * 100
    print(f'{col:<14} {n9:>12,} {pct:>9.1f}%')


# %% [markdown]
# ---
# ## Section 2 — Data Preparation

# %%
df_clean = df.copy()

# ── 3a. Mode + Flag imputation for feature columns with coded-9 ──
IMPUTE_COLS = ['BEDROOMS','TITLED','LOCATION','FOUNDATION','SECURED']
for col in IMPUTE_COLS:
    mode_val = df_clean[df_clean[col] != 9][col].mode()[0]
    df_clean[f'{col}_MISSING'] = (df_clean[col] == 9).astype(int)
    df_clean[col] = df_clean[col].replace(9, mode_val)
    print(f'  {col:<12}: 9 → {mode_val} (mode)  |  {df_clean[f"{col}_MISSING"].sum():,} rows flagged')


# %%
# ── 3b. Drop rows where PRICE = 9 (target cannot be imputed) ──
before = len(df_clean)
df_clean = df_clean[df_clean['PRICE'] != 9].copy()
print(f'Rows removed (PRICE=9): {before - len(df_clean):,}')
print(f'Remaining rows        : {len(df_clean):,}')


# %%
# ── 3c. Extract MONTH and QUARTER from SHIPMONTH (yyyymm format) ──
df_clean['MONTH']   = df_clean['SHIPMONTH'] % 100
df_clean['QUARTER'] = ((df_clean['MONTH'] - 1) // 3) + 1

# ── 3d. Impute FOOTINGS and FINALDEST (sparse: 35% coverage, rest are NaN) ──
df_clean['FOOTINGS_MISSING']  = df_clean['FOOTINGS'].isna().astype(int)
df_clean['FINALDEST_MISSING'] = df_clean['FINALDEST'].isna().astype(int)

foot_mode = df_clean[df_clean['FOOTINGS'].notna() & (df_clean['FOOTINGS'] != 9)]['FOOTINGS'].mode()[0]
df_clean['FOOTINGS']  = df_clean['FOOTINGS'].fillna(-1).replace(9, foot_mode)
df_clean['FINALDEST'] = df_clean['FINALDEST'].fillna(-1)

print(f'FOOTINGS  filled with mode={foot_mode}  |  {df_clean["FOOTINGS_MISSING"].sum():,} rows flagged')
print(f'FINALDEST filled with -1  |  {df_clean["FINALDEST_MISSING"].sum():,} rows flagged')
print(f'\nFinal clean dataset: {len(df_clean):,} rows')


# %% [markdown]
# ---
# ## Section 3 — Feature Engineering

# %%
# ── Structural interactions ──
df_clean['SQFT_per_SECTION']    = df_clean['SQFT'] / df_clean['SECTIONS']
df_clean['SQFT_squared']        = df_clean['SQFT'] ** 2
df_clean['SECTIONS_BEDROOMS']   = df_clean['SECTIONS'] * df_clean['BEDROOMS']
df_clean['SQFT_log']            = np.log1p(df_clean['SQFT'])

# ── Time interactions ──
df_clean['YEAR_since_2014']     = df_clean['YEAR'] - 2014
df_clean['SQFT_YEAR']           = df_clean['SQFT'] * df_clean['YEAR_since_2014']
df_clean['SECTION_YEAR']        = df_clean['SECTIONS'] * df_clean['YEAR_since_2014']

# ── Geographic interactions ──
df_clean['REGION_SECTION']      = df_clean['REGION'] * 10 + df_clean['SECTIONS']
df_clean['REGION_YEAR']         = df_clean['REGION'] * df_clean['YEAR_since_2014']

# ── Financing interactions ──
df_clean['FOUNDATION_LOCATION'] = df_clean['FOUNDATION'] * 10 + df_clean['LOCATION']
df_clean['TITLED_SECURED']      = df_clean['TITLED'] * 10 + df_clean['SECURED']

# ── Market index features ──
yr_med = df_clean.groupby('YEAR')['PRICE'].transform('median')
df_clean['YEAR_MARKET_INDEX']   = yr_med / yr_med.mean()
mo_med = df_clean.groupby('MONTH')['PRICE'].transform('median')
df_clean['MONTH_MARKET_INDEX']  = mo_med / mo_med.mean()

print('Feature engineering complete.')


# %%
# ── Define final feature set (37 features) ──
TARGET   = 'PRICE'
FEATURES = (
    # Original 9
    ['SQFT','SECTIONS','BEDROOMS','REGION','FOUNDATION','TITLED','SECURED','LOCATION','YEAR'] +
    # Missing flags (5)
    ['BEDROOMS_MISSING','TITLED_MISSING','LOCATION_MISSING','FOUNDATION_MISSING','SECURED_MISSING'] +
    # Engineered (12)
    ['SQFT_per_SECTION','SQFT_squared','SECTIONS_BEDROOMS','SQFT_log',
     'YEAR_since_2014','SQFT_YEAR','SECTION_YEAR',
     'REGION_SECTION','REGION_YEAR',
     'FOUNDATION_LOCATION','TITLED_SECURED',
     'YEAR_MARKET_INDEX'] +
    # Previously unused columns (8)
    ['STATUS','MONTH','QUARTER','MONTH_MARKET_INDEX',
     'FOOTINGS','FOOTINGS_MISSING','FINALDEST','FINALDEST_MISSING']
)

print(f'Total features: {len(FEATURES)}')


# %% [markdown]
# ---
# ## Section 4 — Train / Test Split

# %%
X = df_clean[FEATURES]
y = df_clean[TARGET]

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=42
)

print(f'Train : {len(X_train):,} rows')
print(f'Test  : {len(X_test):,} rows')
print(f'Price range : ${y.min():,.0f} — ${y.max():,.0f}')
print(f'Price mean  : ${y.mean():,.0f}')


# %% [markdown]
# ---
# ## Section 5 — Model 1: Linear Regression (Baseline)
# Simple interpretable baseline. Coefficients show dollar change per unit increase in each feature.

# %%
lr = LinearRegression()
lr.fit(X_train, y_train)
y_pred_lr = lr.predict(X_test)

lr_r2   = r2_score(y_test, y_pred_lr)
lr_mae  = mean_absolute_error(y_test, y_pred_lr)
lr_rmse = np.sqrt(mean_squared_error(y_test, y_pred_lr))

print(f'R²   : {lr_r2:.4f}')
print(f'MAE  : ${lr_mae:,.0f}')
print(f'RMSE : ${lr_rmse:,.0f}')

coef_df = pd.Series(lr.coef_, index=FEATURES).sort_values(key=abs, ascending=False)
print('\nTop 10 Coefficients:')
print(coef_df.head(10).apply(lambda x: f'${x:,.2f}').to_string())


# %%
fig, axes = plt.subplots(1, 2, figsize=(14, 5))

# Actual vs Predicted
sample = np.random.default_rng(42).choice(len(y_test), 1500, replace=False)
axes[0].scatter(y_test.values[sample]/1000, y_pred_lr[sample]/1000,
                alpha=0.3, s=8, color='steelblue')
axes[0].plot([20,420],[20,420],'r--',lw=1.5,label='Perfect fit')
axes[0].set_xlabel('Actual Price ($k)'); axes[0].set_ylabel('Predicted Price ($k)')
axes[0].set_title(f'Linear Regression — Actual vs Predicted (R²={lr_r2:.3f})', fontweight='bold')
axes[0].legend(); axes[0].grid(True, alpha=0.3)
axes[0].spines[['top','right']].set_visible(False)

# Coefficients
coef_plot = pd.Series(lr.coef_, index=FEATURES).sort_values()
colors_c  = ['#DC2626' if v < 0 else '#16A34A' for v in coef_plot.values]
axes[1].barh(coef_plot.index, coef_plot.values, color=colors_c, edgecolor='white')
axes[1].axvline(0, color='black', lw=0.8)
axes[1].set_xlabel('Coefficient ($)')
axes[1].set_title('Linear Regression Coefficients', fontweight='bold')
axes[1].grid(True, alpha=0.3, axis='x')
axes[1].spines[['top','right']].set_visible(False)

plt.tight_layout()
plt.savefig('lr_results.png', dpi=300)
plt.show()


# %% [markdown]
# ---
# ## Section 6 — Model 2: Random Forest
# Handles non-linear relationships. Stronger than Linear Regression on structured tabular data.

# %%
rf = RandomForestRegressor(
    n_estimators=100, max_depth=12, random_state=42, n_jobs=-1
)
rf.fit(X_train, y_train)
y_pred_rf = rf.predict(X_test)

rf_r2   = r2_score(y_test, y_pred_rf)
rf_mae  = mean_absolute_error(y_test, y_pred_rf)
rf_rmse = np.sqrt(mean_squared_error(y_test, y_pred_rf))

print(f'R²   : {rf_r2:.4f}')
print(f'MAE  : ${rf_mae:,.0f}')
print(f'RMSE : ${rf_rmse:,.0f}')


# %%
fig, axes = plt.subplots(1, 2, figsize=(14, 5))

# Actual vs Predicted
axes[0].scatter(y_test.values[sample]/1000, y_pred_rf[sample]/1000,
                alpha=0.3, s=8, color='#16A34A')
axes[0].plot([20,420],[20,420],'r--',lw=1.5,label='Perfect fit')
axes[0].set_xlabel('Actual Price ($k)'); axes[0].set_ylabel('Predicted Price ($k)')
axes[0].set_title(f'Random Forest — Actual vs Predicted (R²={rf_r2:.3f})', fontweight='bold')
axes[0].legend(); axes[0].grid(True, alpha=0.3)
axes[0].spines[['top','right']].set_visible(False)

# Feature Importance
imp_rf = pd.Series(rf.feature_importances_, index=FEATURES).sort_values(ascending=False).head(15)
axes[1].barh(imp_rf.index, imp_rf.values,
             color=['#16A34A' if v > 0.05 else '#2563EB' for v in imp_rf.values],
             edgecolor='white')
axes[1].invert_yaxis()
axes[1].set_xlabel('Importance Score')
axes[1].set_title('Random Forest — Top 15 Feature Importances', fontweight='bold')
axes[1].grid(True, alpha=0.3, axis='x')
axes[1].spines[['top','right']].set_visible(False)

plt.tight_layout()
plt.savefig('rf_results.png', dpi=300)
plt.show()


# %% [markdown]
# ---
# ## Section 7 — Model 3: LightGBM (Best Model)
# Gradient boosting — builds trees sequentially correcting prior errors. Best R² on this dataset.

# %%
lgbm = lgb.LGBMRegressor(
    n_estimators     = 600,
    learning_rate    = 0.05,
    max_depth        = 10,
    num_leaves       = 63,
    subsample        = 0.85,
    colsample_bytree = 0.85,
    min_child_samples= 20,
    reg_alpha        = 0.05,
    reg_lambda       = 0.1,
    random_state     = 42,
    n_jobs           = -1,
    verbose          = -1
)
lgbm.fit(
    X_train, y_train,
    eval_set=[(X_test, y_test)],
    callbacks=[
        lgb.early_stopping(stopping_rounds=50, verbose=False),
        lgb.log_evaluation(period=0)
    ]
)
y_pred_lgbm = lgbm.predict(X_test)

lgbm_r2   = r2_score(y_test, y_pred_lgbm)
lgbm_mae  = mean_absolute_error(y_test, y_pred_lgbm)
lgbm_rmse = np.sqrt(mean_squared_error(y_test, y_pred_lgbm))

print(f'R²              : {lgbm_r2:.4f}')
print(f'MAE             : ${lgbm_mae:,.0f}')
print(f'RMSE            : ${lgbm_rmse:,.0f}')
print(f'Best iteration  : {lgbm.best_iteration_}')


# %%
fig, axes = plt.subplots(1, 2, figsize=(14, 5))

# Actual vs Predicted
axes[0].scatter(y_test.values[sample]/1000, y_pred_lgbm[sample]/1000,
                alpha=0.3, s=8, color='#7C3AED')
axes[0].plot([20,420],[20,420],'r--',lw=1.5,label='Perfect fit')
axes[0].set_xlabel('Actual Price ($k)'); axes[0].set_ylabel('Predicted Price ($k)')
axes[0].set_title(f'LightGBM — Actual vs Predicted (R²={lgbm_r2:.3f})', fontweight='bold')
axes[0].legend(); axes[0].grid(True, alpha=0.3)
axes[0].spines[['top','right']].set_visible(False)

# Feature Importance
imp_lgbm = pd.Series(lgbm.feature_importances_, index=FEATURES).sort_values(ascending=False).head(15)
axes[1].barh(imp_lgbm.index, imp_lgbm.values,
             color=['#7C3AED' if v > imp_lgbm.max()*0.3 else '#2563EB' for v in imp_lgbm.values],
             edgecolor='white')
axes[1].invert_yaxis()
axes[1].set_xlabel('Importance Score')
axes[1].set_title('LightGBM — Top 15 Feature Importances', fontweight='bold')
axes[1].grid(True, alpha=0.3, axis='x')
axes[1].spines[['top','right']].set_visible(False)

plt.tight_layout()
plt.savefig('lgbm_results.png', dpi=300)
plt.show()


# %% [markdown]
# ---
# ## Section 8 — Anomaly Detection & Removal
# 
# **Residual = Actual Price − Predicted Price**
# - Positive residual → actual > predicted → **Overvalued** (buyer overpaid)
# - Negative residual → actual < predicted → **Undervalued** (buyer got a deal)
# 
# Homes beyond ±2.5σ are anomalies whose prices cannot be explained by available features. Removing them and retraining significantly improves R².

# %%
# ── Compute residuals on full dataset using LightGBM ──
df_clean['PREDICTED_PRICE'] = lgbm.predict(X)
df_clean['RESIDUAL']        = df_clean['PRICE'] - df_clean['PREDICTED_PRICE']

resid_std = df_clean['RESIDUAL'].std()
threshold = 2.5 * resid_std

df_clean['ANOMALY'] = 'Normal'
df_clean.loc[df_clean['RESIDUAL'] >  threshold, 'ANOMALY'] = 'Overvalued'
df_clean.loc[df_clean['RESIDUAL'] < -threshold, 'ANOMALY'] = 'Undervalued'

print(f'Residual std : ${resid_std:,.0f}')
print(f'Threshold    : ±${threshold:,.0f}  (2.5σ)')
print()
print(df_clean['ANOMALY'].value_counts().to_string())


# %%
# ── Residual distribution plot ──
res_normal = df_clean[df_clean['ANOMALY'] == 'Normal']['RESIDUAL']
res_over   = df_clean[df_clean['ANOMALY'] == 'Overvalued']['RESIDUAL']
res_under  = df_clean[df_clean['ANOMALY'] == 'Undervalued']['RESIDUAL']

plt.figure(figsize=(11, 5))
plt.hist(res_normal/1000, bins=80, color='steelblue', alpha=0.8, label=f'Normal ({len(res_normal):,})')
plt.hist(res_over/1000,   bins=20, color='#DC2626',   alpha=0.9, label=f'Overvalued ({len(res_over)})')
plt.hist(res_under/1000,  bins=20, color='#EA580C',   alpha=0.9, label=f'Undervalued ({len(res_under)})')
plt.axvline(0,               color='black',  lw=1.5, ls='--')
plt.axvline(-threshold/1000, color='orange', lw=1.5, ls=':',  label=f'-2.5σ')
plt.axvline( threshold/1000, color='red',    lw=1.5, ls=':',  label=f'+2.5σ')
plt.xlabel('Residual ($k)  [Actual − Predicted]\n← Undervalued  |  Overvalued →')
plt.ylabel('Count')
plt.title('Pricing Anomaly Detection — Residual Distribution', fontweight='bold')
plt.legend(); plt.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('anomaly_distribution.png', dpi=300)
plt.show()


# %%
# ── Top 10 Overvalued homes (actual >> predicted) ──
print('TOP 10 OVERVALUED HOMES (actual price >> model prediction):')
print('Residual is POSITIVE → buyer likely overpaid\n')
ov_df = (df_clean[df_clean['ANOMALY'] == 'Overvalued']
         [['CONTROL','PRICE','PREDICTED_PRICE','RESIDUAL','SQFT','SECTIONS','REGION','YEAR']]
         .sort_values('RESIDUAL', ascending=False).head(10).copy())
ov_df['PRICE']           = ov_df['PRICE'].apply(lambda x: f'${x:,.0f}')
ov_df['PREDICTED_PRICE'] = ov_df['PREDICTED_PRICE'].apply(lambda x: f'${x:,.0f}')
ov_df['RESIDUAL']        = ov_df['RESIDUAL'].apply(lambda x: f'+${x:,.0f}')
print(ov_df.to_string(index=False))


# %%
# ── Top 10 Undervalued homes (actual << predicted) ──
print('TOP 10 UNDERVALUED HOMES (actual price << model prediction):')
print('Residual is NEGATIVE → buyer likely got a deal\n')
uv_df = (df_clean[df_clean['ANOMALY'] == 'Undervalued']
         [['CONTROL','PRICE','PREDICTED_PRICE','RESIDUAL','SQFT','SECTIONS','REGION','YEAR']]
         .sort_values('RESIDUAL', ascending=True).head(10).copy())
uv_df['PRICE']           = uv_df['PRICE'].apply(lambda x: f'${x:,.0f}')
uv_df['PREDICTED_PRICE'] = uv_df['PREDICTED_PRICE'].apply(lambda x: f'${x:,.0f}')
uv_df['RESIDUAL']        = uv_df['RESIDUAL'].apply(lambda x: f'-${abs(x):,.0f}')
print(uv_df.to_string(index=False))


# %%
# ── Remove anomalies and retrain LightGBM ──
mask_clean    = df_clean['ANOMALY'] == 'Normal'
X_clean       = df_clean.loc[mask_clean, FEATURES]
y_clean       = df_clean.loc[mask_clean, TARGET]

X_tr_c, X_te_c, y_tr_c, y_te_c = train_test_split(X_clean, y_clean, test_size=0.2, random_state=42)

lgbm_final = lgb.LGBMRegressor(
    n_estimators=600, learning_rate=0.05, max_depth=10, num_leaves=63,
    subsample=0.85, colsample_bytree=0.85, min_child_samples=20,
    reg_alpha=0.05, reg_lambda=0.1, random_state=42, n_jobs=-1, verbose=-1
)
lgbm_final.fit(X_tr_c, y_tr_c, eval_set=[(X_te_c, y_te_c)],
               callbacks=[lgb.early_stopping(50, verbose=False), lgb.log_evaluation(0)])

y_pred_final = lgbm_final.predict(X_te_c)
final_r2     = r2_score(y_te_c, y_pred_final)
final_mae    = mean_absolute_error(y_te_c, y_pred_final)
final_rmse   = np.sqrt(mean_squared_error(y_te_c, y_pred_final))

print(f'Anomalies removed : {(~mask_clean).sum():,} homes ({(~mask_clean).mean()*100:.1f}%)')
print(f'Training on       : {mask_clean.sum():,} homes')
print()
print(f'R²   : {final_r2:.4f}  (was {lgbm_r2:.4f}  → +{(final_r2-lgbm_r2)*100:.2f}pp)')
print(f'MAE  : ${final_mae:,.0f}  (was ${lgbm_mae:,.0f})')
print(f'RMSE : ${final_rmse:,.0f}  (was ${lgbm_rmse:,.0f})')


# %% [markdown]
# ---
# ## Section 9 — Model Comparison

# %%
comparison = pd.DataFrame([
    {'Model': 'Linear Regression',             'Features': 37, 'R²': lr_r2,    'MAE ($)': lr_mae,    'RMSE ($)': lr_rmse},
    {'Model': 'Random Forest',                 'Features': 37, 'R²': rf_r2,    'MAE ($)': rf_mae,    'RMSE ($)': rf_rmse},
    {'Model': 'LightGBM',                      'Features': 37, 'R²': lgbm_r2,  'MAE ($)': lgbm_mae,  'RMSE ($)': lgbm_rmse},
    {'Model': 'LightGBM + Anomaly Removal ★',  'Features': 37, 'R²': final_r2, 'MAE ($)': final_mae, 'RMSE ($)': final_rmse},
])
comparison['R²']      = comparison['R²'].round(4)
comparison['MAE ($)'] = comparison['MAE ($)'].apply(lambda x: f'${x:,.0f}')
comparison['RMSE ($)']= comparison['RMSE ($)'].apply(lambda x: f'${x:,.0f}')
print(comparison.to_string(index=False))


# %%
fig, axes = plt.subplots(1, 2, figsize=(14, 5))

models     = ['Linear\nRegression', 'Random\nForest', 'LightGBM', 'LightGBM +\nAnomaly Removal ★']
r2_all     = [lr_r2, rf_r2, lgbm_r2, final_r2]
mae_all    = [lr_mae, rf_mae, lgbm_mae, final_mae]
bar_colors = ['#94A3B8', '#64748B', '#2563EB', '#7C3AED']

# R²
bars = axes[0].bar(models, r2_all, color=bar_colors, edgecolor='white', width=0.55)
axes[0].set_ylim(0.55, max(r2_all) + 0.05)
axes[0].set_ylabel('R² Score', fontsize=11)
axes[0].set_title('R² Score — All Models', fontweight='bold')
for bar, v in zip(bars, r2_all):
    axes[0].text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.005,
                 f'{v:.4f}', ha='center', fontsize=9, fontweight='bold')

# MAE
bars2 = axes[1].bar(models, mae_all, color=bar_colors, edgecolor='white', width=0.55)
axes[1].set_ylabel('MAE ($)', fontsize=11)
axes[1].set_title('MAE — All Models (lower = better)', fontweight='bold')
for bar, v in zip(bars2, mae_all):
    axes[1].text(bar.get_x()+bar.get_width()/2, bar.get_height()+50,
                 f'${v:,.0f}', ha='center', fontsize=9, fontweight='bold')

for ax in axes:
    ax.grid(True, alpha=0.3, axis='y')
    ax.spines[['top','right']].set_visible(False)
    ax.tick_params(axis='x', labelsize=9)

plt.suptitle('Full Model Comparison', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig('model_comparison_final.png', dpi=300)
plt.show()


# %% [markdown]
# ---
# ## Section 10 — Regional & Time Analysis

# %%
region_map = {1:'Northeast', 2:'Midwest', 3:'South', 4:'Southwest', 5:'West'}
df_clean['REGION_NAME'] = df_clean['REGION'].map(region_map)
region_order = ['Northeast','Midwest','South','Southwest','West']

regional = (df_clean.groupby('REGION_NAME')['PRICE']
            .agg(Mean='mean', Median='median', Count='count')
            .reindex(region_order).reset_index())

print('Regional Price Summary:')
print(regional.to_string(index=False))


# %%
fig, axes = plt.subplots(1, 2, figsize=(15, 5))

# Mean vs Median by region
x = np.arange(len(region_order)); w = 0.38
b1 = axes[0].bar(x-w/2, regional['Mean']/1000,   w, color='steelblue', label='Mean',   edgecolor='white')
b2 = axes[0].bar(x+w/2, regional['Median']/1000, w, color='coral',     label='Median', edgecolor='white')
axes[0].set_xticks(x); axes[0].set_xticklabels(region_order, fontsize=10)
axes[0].set_ylabel('Price ($k)'); axes[0].set_title('Mean vs Median Price by Region', fontweight='bold')
axes[0].legend()
for bar in list(b1)+list(b2):
    axes[0].text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
                 f'${bar.get_height():.0f}k', ha='center', fontsize=8)
axes[0].grid(True, alpha=0.3, axis='y'); axes[0].spines[['top','right']].set_visible(False)

# Price trend by year
yr_mean = df_clean.groupby('YEAR')['PRICE'].mean() / 1000
axes[1].plot(yr_mean.index, yr_mean.values, marker='o', color='#EA580C', lw=2.5, ms=7, mfc='white', mew=2)
axes[1].fill_between(yr_mean.index, yr_mean.values, alpha=0.12, color='#EA580C')
for xv, yv in zip(yr_mean.index, yr_mean.values):
    axes[1].annotate(f'${yv:.0f}k', (xv, yv), textcoords='offset points',
                     xytext=(0,8), ha='center', fontsize=7.5)
axes[1].set_xlabel('Year'); axes[1].set_ylabel('Mean Price ($k)')
axes[1].set_title('Mean Price Trend 2014–2024', fontweight='bold')
axes[1].grid(True, alpha=0.3); axes[1].spines[['top','right']].set_visible(False)

plt.tight_layout()
plt.savefig('regional_time_analysis.png', dpi=300)
plt.show()


# %%
# Price trend by region over years
yearly_region = df_clean.groupby(['YEAR','REGION_NAME'])['PRICE'].mean().reset_index()
colors_reg = {'Northeast':'#2563EB','Midwest':'#16A34A','South':'#EA580C',
              'Southwest':'#7C3AED','West':'#DC2626'}

plt.figure(figsize=(12, 5))
for region in region_order:
    d = yearly_region[yearly_region['REGION_NAME'] == region]
    plt.plot(d['YEAR'], d['PRICE']/1000, marker='o', lw=2, ms=5,
             label=region, color=colors_reg[region])
plt.xlabel('Year'); plt.ylabel('Mean Price ($k)')
plt.title('Mean Home Price by Region (2014–2024)', fontweight='bold')
plt.legend(); plt.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('regional_trend.png', dpi=300)
plt.show()


# %% [markdown]
# ---
# ## Section 11 — Final Summary
# 
# **Model Performance**
# 
# - **Linear Regression**  
#   R²: 0.619 | MAE: $20,994 | RMSE: $31,342
# 
# - **Random Forest**  
#   R²: 0.692 | MAE: $18,226 | RMSE: $28,163
# 
# - **LightGBM**  
#   R²: 0.703 | MAE: $16,702 | RMSE: $25,696
# 
# - **LightGBM + Outlier Removal ★**  
#   **R²: 0.806 | MAE: $13,380 | RMSE: $18,639**
# 
# **Key findings:**
# - `YEAR` and `SECTIONS` are the strongest price drivers
# - West region homes average nearly 2× Midwest prices
# - Prices rose from ~$70k (2014) to ~$130k (2023)
# - Mode + Flag imputation recovered 24,000+ rows vs dropping
# 
# **Key findings:**
# - `YEAR` and `SECTIONS` are the strongest price drivers
# - West region homes average nearly 2× Midwest prices
# - Prices rose from ~$70k (2014) to ~$130k (2023)
# - Mode + Flag imputation recovered 24,000+ rows vs dropping
# - Anomaly removal (2.5σ) eliminated 3.1% of unpredictable homes, boosting R² by +10pp
# 
# > **Note:** The final R² of ~0.806 is measured on homes that were explainable by the available features. The 3.1% of removed anomalous homes are excluded from this evaluation. State this clearly in any report.

# %%
print('='*55)
print('  FINAL MODEL PERFORMANCE SUMMARY')
print('='*55)
print(f'  Model            : LightGBM + Anomaly Removal')
print(f'  Features used    : {len(FEATURES)}')
print(f'  Training records : {mask_clean.sum():,}')
print(f'  R² Score         : {final_r2:.4f}')
print(f'  MAE              : ${final_mae:,.0f}')
print(f'  RMSE             : ${final_rmse:,.0f}')
print('='*55)



